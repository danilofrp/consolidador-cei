import sys, os
import numpy as np
import pandas as pd
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from consolidate_cei import consolidate_cei_extracts

def main():
    try:
        action, param = get_args()
    except:
        print(f'\nUsageError: incorrect parameters.\nCorrect usage methods:\npython {__file__} --declaracao [ano]\npython {__file__} --posicao [data:yyyy-mm-dd]')
        return

    transactions = consolidate_cei_extracts(save_to_file = False)

    # transactions = pd.read_excel('consolidado_cei.xlsx', index_col = 'Data', parse_dates = True).sort_index()

    if action == '--declaracao':
        declaration, realised_monthly_stocks, realised_monthly_fii, realised_monthly_options = get_declaration_info(transactions, param)
        realised_monthly = pd.concat([realised_monthly_stocks, realised_monthly_fii, realised_monthly_options], axis = 1).fillna(0).sum(axis = 1)
        realised_monthly.name = "Realizado"
        realised_monthly.index.name = "Mês"

        filename = f'declaracao_{param}.xlsx'
        
        writer = pd.ExcelWriter(filename)
        declaration.to_excel(writer, 'Declaração de Bens')
        realised_monthly.to_excel(writer, 'Lucro Realizado Total')
        realised_monthly_stocks.to_excel(writer, 'Lucro Realizado (Ações)')
        realised_monthly_options.to_excel(writer, 'Lucro Realizado (Opções)')
        realised_monthly_fii.to_excel(writer, 'Lucro Realizado (FIIs)')
        writer.save()

        beutify_positions_excel(filename, history_column = 'F')
        print(f'Declaracao salva em {filename}')

    elif action == '--posicao':
        positions, realised_monthly_stocks, realised_monthly_fii, realised_monthly_options = get_position_info(transactions, param)
        realised_monthly = pd.concat([realised_monthly_stocks, realised_monthly_fii, realised_monthly_options], axis = 1).fillna(0).sum(axis = 1)
        realised_monthly.name = "Realizado"
        realised_monthly.index.name = "Mês"

        filename = f'posicoes_{param}.xlsx'
        positions.to_excel(filename, index = False)

        writer = pd.ExcelWriter(filename)
        positions.to_excel(writer, 'Posições')
        realised_monthly.to_excel(writer, 'Lucro Realizado Total')
        realised_monthly_stocks.to_excel(writer, 'Lucro Realizado (Ações)')
        realised_monthly_options.to_excel(writer, 'Lucro Realizado (Opções)')
        realised_monthly_fii.to_excel(writer, 'Lucro Realizado (FIIs)')
        writer.save()

        beutify_positions_excel(filename, history_column = 'F')
        print(f'Posicoes salvas em {filename}')
    
    return

def get_args():
    action = sys.argv[1]
    if action == '--declaracao':
        try:
            year = int(sys.argv[2])
        except:
            year = date.today().year - 1
        return action, year
    elif action == '--posicao':
        try:
            limit_date = sys.argv[2]
        except:
            limit_date = date.today().strftime('%Y-%m-%d')
        return action, limit_date
    else:
        raise Exception()


def get_declaration_info(transactions, interest_year):
    previous_year = interest_year - 1

    assert len(transactions[:f'{previous_year}-12-31'])>0, f'No transaction found for previus year ({previous_year}). If its your first declaration, use the position option: --position yyy-mm-dd instead'

    positions_previous_year, realised_monthly_stocks_previous_year, realised_monthly_fii_previous_year, realised_monthly_options_previous_year = \
        get_position_info(transactions, f'{previous_year}-12-31')
    positions_interest_year, realised_monthly_stocks_interest_year, realised_monthly_fii_interest_year, realised_monthly_options_interest_year = \
        get_position_info(transactions, f'{interest_year}-12-31', ignore_history_previous_to = interest_year)

    declaration = prepare_declaration_dataframe(positions_previous_year, positions_interest_year, interest_year)
    return declaration, realised_monthly_stocks_interest_year, realised_monthly_fii_interest_year, realised_monthly_options_interest_year


def prepare_declaration_dataframe(positions_previous_year, positions_interest_year, interest_year):
    previous_year = interest_year - 1
    previous_year_declaration = pd.DataFrame()
    previous_year_declaration[f'Posição em {previous_year}-12-31'] = positions_previous_year['Valor Total']

    interest_year_declaration = pd.DataFrame()
    interest_year_declaration[f'Posição em {interest_year}-12-31'] = positions_interest_year['Valor Total']
    interest_year_declaration[f'Quantidade em {interest_year}-12-31'] = positions_interest_year['Quantidade']
    interest_year_declaration[f'Preço médio em {interest_year}-12-31'] = positions_interest_year['Preço Médio']
    interest_year_declaration[f'Histórico em {interest_year}-12-31'] = positions_interest_year['Historico']

    declaration = pd.merge(previous_year_declaration, interest_year_declaration, left_index = True, right_index = True, how = 'outer')
    declaration.sort_index(inplace = True)
    declaration.fillna(0, inplace = True)
    declaration = declaration[(declaration[f'Posição em {previous_year}-12-31'] != 0) | (declaration[f'Posição em {interest_year}-12-31'] != 0)]
    declaration.index.name = 'Ativo'

    return declaration


def get_position_info(transactions, limit_date, ignore_history_previous_to = 1900):
    ignore_history_previous_to = int(ignore_history_previous_to)

    vencimentos_opcoes = pd.Series(name = "qtd", index = pd.MultiIndex.from_arrays([[], []], names = ["date", "code"]), dtype = "float64")

    for transaction_date, transaction in transactions[:limit_date].iterrows():
        if transaction['Tipo'] == "Opção":
            try:
                vencimentos_opcoes.loc[(transaction["Prazo"], transaction['Codigo'])] += transaction["Quantidade"]
            except KeyError: 
                vencimentos_opcoes.loc[(transaction["Prazo"], transaction['Codigo'])] = transaction["Quantidade"]
    vencimentos_opcoes = vencimentos_opcoes.replace(0, np.nan).dropna()

    for (transaction_date, codigo), quantidade in vencimentos_opcoes.iteritems():
        new_transaction = pd.Series({
            "Fluxo": "V" if quantidade > 0 else "C",
            "Mercado": "Opção (Vencimento)",
            "Prazo": np.nan,
            "Codigo": codigo,
            "Ativo": "",
            "Quantidade": -quantidade,
            "Preco": 0,
            "Valor Total": 0, 
            "Corretora": "",
            "Tipo": "Opção (Vencimento)",
        }, name = transaction_date)
        transactions = transactions.append(new_transaction)
    transactions = transactions.sort_index()

    positions = {}
    realised_monthly_stocks = pd.Series(name = 'Realizado', dtype = "float64")
    realised_monthly_fii = pd.Series(name = 'Realizado', dtype = "float64")
    realised_monthly_options = pd.Series(name = 'Realizado', dtype = "float64")        

    for transaction_date, transaction in transactions[:limit_date].iterrows():

        codigo = transaction['Codigo']
        
        if codigo not in positions:
            position = {
                'asset': codigo, 
                'qtd': 0,
                'preco_medio': 0,
                'status': None,
                'historico': []
            }
        else:
            position = positions[codigo]

        vencimento_opcao = (transaction["Tipo"] == "Opção (Vencimento)")
        ignore_history = ignore_history_previous_to > transaction_date.year
        if transaction['Fluxo'] == 'C':
            position, realised = process_buy(transaction_date, position, transaction, ignore_history, vencimento_opcao)
        elif transaction['Fluxo'] == 'V':
            position, realised = process_sell(transaction_date, position, transaction, ignore_history, vencimento_opcao)

        position = update_position_status(position)

        if transaction['Tipo'] == "Ação":
            realised_monthly_stocks = realised_monthly_stocks.append(
                    pd.Series([realised], name = "Realizado Mensal", index = [transaction_date])
                )
        elif transaction['Tipo'] == "FII":
            realised_monthly_fii = realised_monthly_fii.append(
                pd.Series([realised], name = "Realizado Mensal", index = [transaction_date])
            )
        elif transaction['Tipo'] in ("Opção", "Opção (Vencimento)"):
            realised_monthly_options = realised_monthly_options.append(
                pd.Series([realised], name = "Realizado Mensal", index = [transaction_date])
            )

            try:
                vencimentos_opcoes.loc[(transaction["Prazo"], codigo)] += transaction["Quantidade"]
            except KeyError: 
                vencimentos_opcoes.loc[(transaction["Prazo"], codigo)] = transaction["Quantidade"]
            vencimentos_opcoes = vencimentos_opcoes.replace(0, np.nan).dropna()
        
        positions[codigo] = position


    if(len(realised_monthly_stocks)>0):
        realised_monthly_stocks = realised_monthly_stocks.groupby(pd.Grouper(freq = "M")).sum().resample("M").asfreq().fillna(0)
        realised_monthly_stocks.index = realised_monthly_stocks.index.date
        realised_monthly_stocks.index.name = "Mês"
        realised_monthly_stocks.name = "Realizado"

    if(len(realised_monthly_fii)>0):
        realised_monthly_fii = realised_monthly_fii.groupby(pd.Grouper(freq = "M")).sum().resample("M").asfreq().fillna(0)
        realised_monthly_fii.index = realised_monthly_fii.index.date
        realised_monthly_fii.index.name = "Mês"
        realised_monthly_fii.name = "Realizado"

    if(len(realised_monthly_options)>0):
        realised_monthly_options = realised_monthly_options.groupby(pd.Grouper(freq = "M")).sum().resample("M").asfreq().fillna(0)
        realised_monthly_options.index = realised_monthly_options.index.date
        realised_monthly_options.index.name = "Mês"
        realised_monthly_options.name = "Realizado"

    positions_df = prepare_position_dataframe(positions)
    return positions_df, realised_monthly_stocks.round(2), realised_monthly_fii.round(2), realised_monthly_options.round(2)


def process_buy(transaction_date, position, transaction, ignore_history = False, vencimento_opcao = False):
    realised = 0
    if position['status'] != 'short':
        position['preco_medio'] = (position['qtd']*position['preco_medio'] + transaction['Quantidade']*transaction['Preco'])/(position['qtd'] + transaction['Quantidade'])
    else:
        realised = transaction['Quantidade'] * (position['preco_medio'] - transaction['Preco'])
    position['qtd'] = position['qtd'] + transaction['Quantidade']
    if not ignore_history:
        if vencimento_opcao:
            position['historico'].append(f'{transaction_date.date()} Vencimento de {position["asset"]} ({transaction["Quantidade"]} x {transaction["Preco"]:.2f})')
        else:
            position['historico'].append(f'{transaction_date.date()} Compra de {position["asset"]} ({transaction["Quantidade"]} x {transaction["Preco"]:.2f})')

    return position, realised


def process_sell(transaction_date, position, transaction, ignore_history = False, vencimento_opcao = False):
    realised = 0
    if position['status'] == 'long':
        realised = (-transaction['Quantidade']) * (transaction['Preco'] - position['preco_medio'])
    else:
        position['preco_medio'] = ((-position['qtd'])*position['preco_medio'] - transaction['Quantidade']*transaction['Preco'])/((-position['qtd']) - transaction['Quantidade'])
    position['qtd'] = position['qtd'] + transaction['Quantidade']
    if not ignore_history:
        if vencimento_opcao:
            position['historico'].append(f'{transaction_date.date()} Vencimento de {position["asset"]} ({-transaction["Quantidade"]} x {transaction["Preco"]:.2f})')
        else:
            position['historico'].append(f'{transaction_date.date()} Venda de {position["asset"]} ({-transaction["Quantidade"]} x {transaction["Preco"]:.2f})')

    return position, realised


def update_position_status(position):
    if position['qtd'] == 0:
        position['preco_medio'] = 0
        position['status'] = None
    elif position['qtd'] > 0:
        position['status'] = 'long'
    elif position['qtd'] < 0:
        position['status'] = 'short'
    return position


def prepare_position_dataframe(positions):
    positions_df = pd.DataFrame(positions).T.sort_index()
    positions_df["preco_medio"] = positions_df["preco_medio"].astype(float).round(2)
    positions_df['Valor Total'] = positions_df['preco_medio'] * positions_df['qtd']
    positions_df.rename(columns = {'asset': 'Ativo', 'preco_medio': 'Preço Médio', 'qtd': 'Quantidade'}, inplace = True)
    positions_df['Historico'] = positions_df['historico'].apply(lambda x: str.join('\n', x))
    positions_df = positions_df[['Ativo', 'status', 'Preço Médio', 'Quantidade', 'Valor Total', 'Historico']]
    return positions_df


def beutify_positions_excel(filename, history_column):
    wb = load_workbook(filename)
    ws = wb.active
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions[history_column].width = 50
    for cell in ws[history_column]:
        cell.alignment = Alignment(wrapText=True)
    wb.save(filename)



if __name__ == "__main__":
    main()